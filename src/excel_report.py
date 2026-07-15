"""
excel_report.py

Builds the final formatted Excel report (reports/Finance_Report.xlsx)
containing a Summary sheet, the cleaned Transactions, Monthly Analysis,
Categories breakdown, and an embedded Charts sheet.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analyzer import AnalysisResult

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(bold=True, size=12, color="1F4E78")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)
CURRENCY_FORMAT = "#,##0.00"
PERCENT_FORMAT = "0.0%"

RED_FILL = PatternFill(start_color="F8CBCB", end_color="F8CBCB", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _style_header_row(ws: Worksheet, row: int, num_columns: int) -> None:
    for col_idx in range(1, num_columns + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws: Worksheet, num_columns: int, min_width: int = 10, max_width: int = 45) -> None:
    for col_idx in range(1, num_columns + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def _write_dataframe(ws: Worksheet, df: pd.DataFrame, start_row: int = 1,
                      currency_columns: list[str] | None = None) -> int:
    """Write a DataFrame starting at start_row with header styling and borders.
    Returns the row number after the last written row."""
    currency_columns = currency_columns or []
    headers = list(df.columns)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header_row(ws, start_row, len(headers))

    for row_offset, (_, row) in enumerate(df.iterrows(), start=1):
        row_number = start_row + row_offset
        for col_idx, header in enumerate(headers, start=1):
            value = row[header]
            if pd.isna(value):
                value = ""
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            cell = ws.cell(row=row_number, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if header in currency_columns:
                cell.number_format = CURRENCY_FORMAT

    _autosize_columns(ws, len(headers))
    return start_row + len(df) + 1


def _build_summary_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Summary")
    overall = analysis.overall
    categories = analysis.categories
    payments = analysis.payments

    ws["A1"] = "Finance Analytics — Summary Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    kpi_rows = [
        ("Total Income", overall.total_income),
        ("Total Expense", overall.total_expense),
        ("Net Profit", overall.net_profit),
        ("Average Expense", overall.average_expense),
        ("Average Income", overall.average_income),
        ("Maximum Expense", overall.max_expense),
        ("Maximum Income", overall.max_income),
        ("Minimum Expense", overall.min_expense),
        ("Minimum Income", overall.min_income),
        ("Total Transactions", overall.total_transactions),
        ("Total Expense Count", overall.total_expense_count),
        ("Total Income Count", overall.total_income_count),
        ("Most Expensive Category", categories.most_expensive_category),
        ("Most Profitable Category", categories.most_profitable_category),
        ("Most Used Payment Method", payments.most_used_method),
    ]

    ws["A3"] = "KPI"
    ws["B3"] = "Value"
    _style_header_row(ws, 3, 2)

    money_labels = {
        "Total Income", "Total Expense", "Net Profit", "Average Expense",
        "Average Income", "Maximum Expense", "Maximum Income",
        "Minimum Expense", "Minimum Income",
    }

    row = 4
    for label, value in kpi_rows:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = THIN_BORDER
        if label in money_labels:
            cell.number_format = CURRENCY_FORMAT
        row += 1

    # Conditional formatting: highlight Net Profit red if negative, green if positive.
    net_profit_row = 4 + [label for label, _ in kpi_rows].index("Net Profit")
    ws.conditional_formatting.add(
        f"B{net_profit_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"B{net_profit_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL),
    )

    _autosize_columns(ws, 2)


def _build_transactions_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Transactions")
    display_df = df.copy()
    end_row = _write_dataframe(ws, display_df, start_row=1, currency_columns=["Amount"])

    # Conditional formatting on the Amount column: red for large expenses,
    # green for income rows.
    amount_col_idx = list(display_df.columns).index("Amount") + 1
    type_col_idx = list(display_df.columns).index("Type") + 1
    amount_letter = get_column_letter(amount_col_idx)
    last_data_row = end_row - 1

    if last_data_row >= 2:
        expense_threshold = float(display_df.loc[display_df["Type"] == "Expense", "Amount"].quantile(0.9)) \
            if (display_df["Type"] == "Expense").any() else 0

        ws.conditional_formatting.add(
            f"{amount_letter}2:{amount_letter}{last_data_row}",
            CellIsRule(operator="greaterThan", formula=[str(expense_threshold)], fill=RED_FILL),
        )

    ws.freeze_panes = "A2"


def _build_monthly_analysis_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Monthly Analysis")
    monthly = analysis.monthly

    df = pd.DataFrame({
        "Month": monthly.monthly_income.index,
        "Income": monthly.monthly_income.values,
        "Expense": monthly.monthly_expense.values,
        "Profit": monthly.monthly_profit.values,
        "Income Growth %": monthly.income_growth_rate.values,
        "Expense Growth %": monthly.expense_growth_rate.values,
    })

    end_row = _write_dataframe(ws, df, start_row=1, currency_columns=["Income", "Expense", "Profit"])

    growth_income_col = get_column_letter(list(df.columns).index("Income Growth %") + 1)
    growth_expense_col = get_column_letter(list(df.columns).index("Expense Growth %") + 1)
    last_data_row = end_row - 1
    for row_idx in range(2, last_data_row + 1):
        ws[f"{growth_income_col}{row_idx}"].number_format = "0.0\"%\""
        ws[f"{growth_expense_col}{row_idx}"].number_format = "0.0\"%\""

    ws.freeze_panes = "A2"


def _build_categories_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Categories")
    categories = analysis.categories

    ws["A1"] = "Top Expense Categories"
    ws["A1"].font = SUBTITLE_FONT
    expense_df = pd.DataFrame({
        "Category": categories.top_expense_categories.index,
        "Total Amount": categories.top_expense_categories.values,
        "Percent of Total": (categories.expense_category_percent
                              .reindex(categories.top_expense_categories.index).values) / 100,
    })
    next_row = _write_dataframe(ws, expense_df, start_row=2, currency_columns=["Total Amount"])
    percent_col = get_column_letter(list(expense_df.columns).index("Percent of Total") + 1)
    for row_idx in range(3, next_row):
        ws[f"{percent_col}{row_idx}"].number_format = PERCENT_FORMAT

    start_income_row = next_row + 2
    ws.cell(row=start_income_row, column=1, value="Top Income Categories").font = SUBTITLE_FONT
    income_df = pd.DataFrame({
        "Category": categories.top_income_categories.index,
        "Total Amount": categories.top_income_categories.values,
        "Percent of Total": (categories.income_category_percent
                              .reindex(categories.top_income_categories.index).values) / 100,
    })
    final_row = _write_dataframe(ws, income_df, start_row=start_income_row + 1,
                                  currency_columns=["Total Amount"])
    percent_col2 = get_column_letter(list(income_df.columns).index("Percent of Total") + 1)
    for row_idx in range(start_income_row + 2, final_row):
        ws[f"{percent_col2}{row_idx}"].number_format = PERCENT_FORMAT

    _autosize_columns(ws, 3)


def _build_charts_sheet(wb: Workbook, chart_paths: list[Path]) -> None:
    ws = wb.create_sheet("Charts")
    ws["A1"] = "Generated Charts"
    ws["A1"].font = TITLE_FONT

    row_cursor = 3
    for chart_path in chart_paths:
        if not chart_path.exists():
            continue
        ws.cell(row=row_cursor, column=1, value=chart_path.stem.replace("_", " ").title()).font = SUBTITLE_FONT
        image = XLImage(str(chart_path))
        # Scale down large PNGs so they fit nicely in the sheet.
        max_width_px = 640
        if image.width > max_width_px:
            scale = max_width_px / image.width
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
        anchor_row = row_cursor + 1
        ws.add_image(image, f"A{anchor_row}")
        row_cursor = anchor_row + int(image.height / 20) + 3

    ws.column_dimensions["A"].width = 90


def build_excel_report(df: pd.DataFrame, analysis: AnalysisResult,
                        chart_paths: list[Path], output_path: Path) -> Path:
    """
    Build the full formatted Excel report and save it to output_path.

    Args:
        df: Cleaned transactions DataFrame.
        analysis: Result of run_full_analysis().
        chart_paths: Paths to previously generated chart PNGs.
        output_path: Destination .xlsx path.

    Returns:
        The path the report was saved to.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _build_summary_sheet(wb, analysis)
    _build_transactions_sheet(wb, df)
    _build_monthly_analysis_sheet(wb, analysis)
    _build_categories_sheet(wb, analysis)
    _build_charts_sheet(wb, chart_paths)

    wb.save(output_path)
    logger.info("Excel report saved to %s", output_path)
    return output_path
