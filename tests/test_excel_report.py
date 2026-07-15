"""Tests for src/excel_report.py"""

import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

SRC_DIR = Path(__file__).resolve().parent.parent / "src"
sys.path.insert(0, str(SRC_DIR))

from analyzer import run_full_analysis  # noqa: E402
from excel_report import build_excel_report  # noqa: E402


@pytest.fixture
def sample_df() -> pd.DataFrame:
    return pd.DataFrame({
        "Date": pd.to_datetime([
            "2026-01-05", "2026-01-10", "2026-02-05", "2026-02-15", "2026-02-20",
        ]),
        "Type": ["Income", "Expense", "Income", "Expense", "Expense"],
        "Category": ["Salary", "Food", "Salary", "Rent", "Food"],
        "Description": ["Salary", "Lunch", "Salary", "Rent", "Dinner"],
        "Amount": [400000, 5000, 420000, 150000, 7000],
        "Payment Method": ["Bank", "Card", "Bank", "Bank", "Cash"],
        "Month": ["January", "January", "February", "February", "February"],
    })


def test_build_excel_report_creates_file(tmp_path: Path, sample_df: pd.DataFrame) -> None:
    analysis = run_full_analysis(sample_df)
    output_path = tmp_path / "Finance_Report.xlsx"
    result_path = build_excel_report(sample_df, analysis, [], output_path)
    assert result_path.exists()


def test_build_excel_report_has_all_sheets(tmp_path: Path, sample_df: pd.DataFrame) -> None:
    analysis = run_full_analysis(sample_df)
    output_path = tmp_path / "Finance_Report.xlsx"
    build_excel_report(sample_df, analysis, [], output_path)

    wb = load_workbook(output_path)
    expected_sheets = {"Summary", "Transactions", "Monthly Analysis", "Categories", "Charts"}
    assert expected_sheets.issubset(set(wb.sheetnames))


def test_build_excel_report_transactions_row_count(tmp_path: Path, sample_df: pd.DataFrame) -> None:
    analysis = run_full_analysis(sample_df)
    output_path = tmp_path / "Finance_Report.xlsx"
    build_excel_report(sample_df, analysis, [], output_path)

    wb = load_workbook(output_path)
    ws = wb["Transactions"]
    # header row + number of data rows
    assert ws.max_row == len(sample_df) + 1
